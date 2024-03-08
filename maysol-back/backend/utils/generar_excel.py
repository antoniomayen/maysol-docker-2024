# coding=utf-8
import os
import datetime
from openpyxl import load_workbook, Workbook, drawing
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from django.conf import settings
from django.http import HttpResponse
from backend.models import Proyecto, Categorias, CambioMoneda
from backend.utils.meses import MESES
from backend.utils.categorias import CF, PL


def descargar_reporte(datos, params):
    """
    Funcion para generar el excel de reporte de categorias
    :return: url donde debe descargar el archivo recien generado
    """

    CF = datos['CF']
    _PL = datos['PL']
    deudas = datos['deudas']
    ventas = datos['ventas']
    ventas_totales = datos['ventas_totales']
    ventas_ateriores = datos['ventas_ateriores']
    moneda = datos['moneda']
    recuperacion = datos['recuperacion']

    idioma = params.get('idioma', 'es')
    template_name = 'template-es.xlsx' if idioma == 'es' else 'template-jp.xlsx'
    template_path = os.path.join(settings.BASE_DIR, template_name)
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = "ReporteCL_PL"
    anio = params['anio']

    empresa_id = params.get('empresa', None)
    empresa = ''
    try:
        _empresa = Proyecto.objects.get(pk=empresa_id)
        empresa = _empresa.nombre
    except:
        empresa = ''
        pass

    # Encabezado
    ws['B1'] = 'Moneda: ' + moneda
    ws['D1'] = empresa + '(' + anio + ')'
    ws['K1'] = anio

    # Textos anios anteriores
    ws['B60'] = str(anio) + '年度売上高PL実績（税別)' if idioma == 'jp' else 'ventas totales PL' + str(anio)
    ws['B61'] = str(int(anio) - 1) + '年度売上高PL実績(税別)' if idioma == 'jp' else 'ventas totales PL' + str(int(anio) - 1)
    ws['B62'] = '営業利益　対' + str(int(anio) - 1) + '年度比' if idioma == 'jp' else 'Utilidad operativa vs. fiscal' + str(int(anio) - 1)
    ws['B63'] = str(anio) + '年度営業利益PL実績' if idioma == 'jp' else 'Utilidades totales PL ' + str(anio)
    ws['B64'] = str(int(anio) - 1) + '年度営業利益PL実績' if idioma == 'jp' else 'Utilidades totales PL ' + str(int(anio) - 1)
    ws['B65'] = '対' + str(anio) + '年度　売上予算比' if idioma == 'jp' else 'FY'+ str(anio) +' relación de presupuesto de ventas'
    ws['B66'] = str(anio) + '年度売上高PL実績（税別)' if idioma == 'jp' else 'ventas totales PL ' + str(anio)
    ws['B67'] = str(anio) + '年度必達予算（税別)' if idioma == 'jp' else 'Meta de ventas ' + str(anio)

    # Anio fiscal
    ws['S7'] = str(anio) + '年度通期' if idioma == 'jp' else 'Año completo ' + str(anio)
    ws['S56'] = str(anio) + '年度通期' if idioma == 'jp' else 'Año completo ' + str(anio)
    ws['G6'] = str(anio) + '年度残高' if idioma == 'jp' else 'Balance del año ' + str(anio)

    # Cuartos
    ws['U7'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['V7'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['W7'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['X7'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)

    ws['U56'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['V56'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['W56'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)
    ws['X56'] = str(anio) + '年度' if idioma == 'jp' else 'Año' + str(anio)

    row_cf = 11  # Fila donde comienzan las categorias de CF
    row_pl = 84  # Fila donde comienzan las categorias de PL

    # contiene el numero de la fila en la que se encuentra un titulo de PL y este ya esta en el template
    titulos_pl = [92, 93, 97, 98, 101, 102, 104, 105, 107, 108, 110, 111, 113, 114, 117, 118, 123,
                  124, 131, 132, 135, 136, 139, 140, 145, 146, 151, 152, 157, 158]
    # Contiene el numero de las filas que tienen que quedar en blanco
    blancos_pl = [91, 96, 100, 103, 106, 109, 112, 116, 122, 130, 134, 138, 144, 156, 162]

    # contiene el numero de la fila en la que se encuentra un titulo de CF y este ya esta en el template
    # La fila 11 esta porque los datos de ventas viene por los pagos
    titulos_cf = [11, 16, 23, 29, 34, 36, 41]

    # Recorre cada categoria de CF
    for res in CF:
        while titulos_cf.__contains__(row_cf):
            row_cf += 1
        ws['B' + str(row_cf)] = CF[res]['es'] if idioma == 'es' else CF[res]['jp']
        ws['D' + str(row_cf)] = CF[res]['total_mes']

        meses = CF[res]['meses']
        for _mes in meses:
            ws[str(MESES[str(_mes)]) + str(row_cf)] = meses[_mes]['total']
        row_cf += 1

    # Recorre cada categoria de PL
    for pl in _PL:
        while titulos_pl.__contains__(row_pl) or blancos_pl.__contains__(row_pl):
            row_pl += 1
        ws['B' + str(row_pl)] = _PL[pl]['es'] if idioma == 'es' else _PL[pl]['jp']
        ws['E' + str(row_pl)] = _PL[pl]['total_mes']

        meses = _PL[pl]['meses']
        for _mes in meses:
            ws[str(MESES[str(_mes)]) + str(row_pl)] = meses[_mes]['total']

        row_pl += 1

    ws['D15'] = deudas['haber']['total']  # Coloca las deudas cobradas, fila 15
    ws['D22'] = deudas['debe']['total']  # Coloca las deudas pagadas, fila 22
    ws['D70'] = ventas['total']  # Coloca el total de las ventas, fila 70
    ws['D61'] = ventas_ateriores['total']  # Coloca el total de las ventas del año pasado, fila 61

    for mes in MESES:
        ws[MESES[mes] + '15'] = deudas['haber'][mes]  # deudas cobradas, fila 15
        ws[MESES[mes] + '22'] = deudas['debe'][mes]  # deudas pagadas, fila 22
        ws[MESES[mes] + '35'] = deudas['sumatoria'][mes]  # deudas pagadas y cobradas, fila 35
        ws[MESES[mes] + '11'] = ventas[mes]  # ventas (pagos), fila 11
        ws[MESES[mes] + '61'] = ventas_ateriores[mes]  # ventas anteriores (pagos), fila 61
        ws[MESES[mes] + '70'] = ventas_totales[mes]  # ventas tomando el monto total dela OV, fila 70
        ws[MESES[mes] + '159'] = recuperacion['159'][mes]  # Depresiacion, fila 159
        ws[MESES[mes] + '160'] = recuperacion['160'][mes]  # Depreasiacion de sistemas, fila 160
        ws[MESES[mes] + '161'] = recuperacion['161'][mes]  # 設立投資 - Establecimiento de inversión, fila 161

    response = HttpResponse(content_type='application/ms-excel')
    filename = 'Reporte.xlsx'
    response['Content-Disposition'] = 'attachment; filename='+filename+''
    wb.save(response)
    return response

def descargar_reporte_detalle(datos, params):
    """
    Funcion para generar el excel de reporte de categorias
    :return: url donde debe descargar el archivo recien generado
    """
    cambio_moneda=CambioMoneda.objects.last()
    template_name = 'template-detalle.xlsx'
    template_path = os.path.join(settings.BASE_DIR, template_name)
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = f"Reporte CF detalle"

    
    ws['A2'] = f"{datos['empresa']}"
    ws['C3'] = f"{datos['anio']}"
    ws['D3'] = f"{datos['mes']}"

    for index, item in enumerate(datos['detalle']):
        if item.get('categoria') is not None:
            categoria=Categorias.objects.get(pk=item.get('categoria'))
            if datos['tipo'] == "CF":
                categoria = CF[categoria.cf]
                n_categoria=categoria['es'] + " - " + categoria['jp']
            else:
                categoria = PL[categoria.pl]
                n_categoria=categoria['es'] + " - " + categoria['jp']
        else:
            categoria=None
        ws[f'A{index + 8}'].alignment = Alignment(horizontal='center')
        ws[f'C{index + 8}'].alignment = Alignment(horizontal='center')
        ws[f'D{index + 8}'].alignment = Alignment(horizontal='right')
        ws[f'E{index + 8}'].alignment = Alignment(horizontal='center')
        
        if item.get('anulado')== True:
            ws[f'A{index + 8}'].font = Font(color=colors.RED)
            ws[f'B{index + 8}'].font = Font(color=colors.RED)
            ws[f'C{index + 8}'].font = Font(color=colors.RED)
            ws[f'D{index + 8}'].font = Font(color=colors.RED)
            ws[f'E{index + 8}'].font = Font(color=colors.RED)
            ws[f'F{index + 8}'].font = Font(color=colors.RED)

        ws[f'A{index + 8}'] = f"{datetime.datetime.strptime(item.get('fecha'), '%Y-%m-%d').strftime('%d-%m-%Y')}"
        ws[f'B{index + 8}'] = f"{item.get('concepto')}"
        ws[f'C{index + 8}'] = f"{item.get('destino')}"

        if datos['moneda']=="GTQ":
            ws[f'D{index + 8}'] = f"Q{round(float(item.get('monto')),2)}"
        elif datos['moneda']=="USD":
            ws[f'D{index + 8}'] = f"${round(float(item.get('monto') / cambio_moneda.cambio_dolar),2)}"
        elif datos['moneda']=="YEN":
            ws[f'D{index + 8}'] = f"¥{round(float(item.get('monto') * cambio_moneda.cambio_yen),2)}"
        ws[f'E{index + 8}'] = f"{n_categoria if categoria is not None else ''}"
        ws[f'F{index + 8}'] = f"{item.get('comentario') if item.get('comentario') is not None else ''}"
        

    response = HttpResponse(content_type='application/ms-excel')
    filename = f"Reporte CF detalle.xlsx"
    response['Content-Disposition'] = 'attachment; filename='+filename+''
    wb.save(response)
    return response
