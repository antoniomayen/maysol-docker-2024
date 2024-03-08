# coding=utf-8
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from django.conf import settings
from django.http import HttpResponse

def descargar_reporte_ventas(datos, suma, nombre, fechaInicio, fechaFinal):
    """
    Funcion para generar el excel de reporte de categorias
    :return: url donde debe descargar el archivo recien generado
    """

    template_name = 'template-ventas.xlsx'
    template_path = os.path.join(settings.BASE_DIR, template_name)
    wb = load_workbook(template_path)
    ws = wb.active
    ws.title = f"Reporte Cobros Ventas - {nombre}"

    ws['C3'] = f"Q{float(suma)}"
    ws['A2'] = f"{fechaInicio.strftime('%d-%m-%Y')} - {fechaFinal.strftime('%d-%m-%Y')}"
    ws['A5'] = f"{nombre}"

    for index, item in enumerate(datos):
        ws[f'A{index + 8}'].alignment = Alignment(horizontal='center')
        ws[f'D{index + 8}'].alignment = Alignment(horizontal='right')
        ws[f'E{index + 8}'].alignment = Alignment(horizontal='center')

        ws[f'A{index + 8}'] = f"{datetime.datetime.strptime(item.get('fecha'), '%Y-%m-%d').strftime('%d-%m-%Y')}"
        ws[f'B{index + 8}'] = f"{item.get('cliente')}"
        ws[f'C{index + 8}'] = f"{item.get('usuario')}"
        ws[f'D{index + 8}'] = f"Q{float(item.get('monto'))}"
        ws[f'E{index + 8}'] = f"{item.get('boleta') if item.get('boleta') is not None else ''}"
        ws[f'F{index + 8}'] = f"{item.get('concepto')}"
        ws[f'G{index + 8}'] = f"{item.get('empresa')}"

    response = HttpResponse(content_type='application/ms-excel')
    filename = f"Reporte Cobros Ventas.xlsx"
    response['Content-Disposition'] = 'attachment; filename='+filename+''
    wb.save(response)
    return response
